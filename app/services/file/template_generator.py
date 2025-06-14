"""
模板生成器
"""
import io
from typing import Dict, List, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from app.services.dify.models import WorkflowParameters, WorkflowParameter, WorkflowParameterType
from app.core.exceptions import FileProcessingException
from app.core.logging import get_logger

logger = get_logger(__name__)


class TemplateGenerator:
    """模板生成器"""
    
    def __init__(self):
        """初始化模板生成器"""
        self.default_column_width = 20
        self.max_column_width = 50
        self.min_column_width = 10
        
        # 预定义样式
        self._setup_styles()
    
    def _setup_styles(self):
        """设置预定义样式"""
        # 表头样式
        self.header_style = {
            "font": Font(bold=True, color="FFFFFF", size=12),
            "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # 描述行样式
        self.description_style = {
            "font": Font(italic=True, color="666666", size=10),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # 示例行样式
        self.example_style = {
            "font": Font(color="999999", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # 标题样式
        self.title_style = {
            "font": Font(bold=True, size=14, color="2F4F4F"),
            "alignment": Alignment(horizontal="left", vertical="center")
        }
        
        # 子标题样式
        self.subtitle_style = {
            "font": Font(bold=True, size=12, color="4682B4"),
            "alignment": Alignment(horizontal="left", vertical="center")
        }
    
    def generate_workflow_template(self, parameters: WorkflowParameters) -> io.BytesIO:
        """
        生成工作流Excel模板
        
        Args:
            parameters: 工作流参数信息
            
        Returns:
            Excel文件的字节流
        """
        try:
            logger.info(f"开始生成工作流模板: {parameters.workflow_name}")
            
            # 创建工作簿
            wb = Workbook()
            
            # 生成数据工作表
            self._create_data_sheet(wb, parameters)
            
            # 生成说明工作表
            self._create_instructions_sheet(wb, parameters)
            
            # 生成参数详情工作表
            self._create_parameters_sheet(wb, parameters)
            
            # 设置活动工作表为数据表
            wb.active = wb["批量数据"]
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"工作流模板生成完成: {parameters.workflow_name}")
            return output
            
        except Exception as e:
            logger.error(f"生成工作流模板失败: {str(e)}")
            raise FileProcessingException(f"生成工作流模板失败: {str(e)}")
    
    def _create_data_sheet(self, wb: Workbook, parameters: WorkflowParameters):
        """创建数据工作表"""
        # 删除默认工作表并创建新的
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        ws = wb.create_sheet("批量数据", 0)
        
        # 生成列信息
        columns_info = self._prepare_columns_info(parameters)
        
        # 写入表头行
        self._write_header_row(ws, columns_info, 1)
        
        # 写入描述行
        self._write_description_row(ws, columns_info, 2)
        
        # 写入示例行
        self._write_example_row(ws, columns_info, 3)
        
        # 调整列宽
        self._adjust_column_widths(ws, columns_info)
        
        # 冻结前三行
        ws.freeze_panes = "A4"
        
        # 设置打印选项
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    def _prepare_columns_info(self, parameters: WorkflowParameters) -> List[Dict[str, Any]]:
        """准备列信息"""
        columns_info = []
        
        # 添加参数列
        for param in parameters.parameters:
            column_info = {
                "name": param.name,
                "description": param.description or f"{param.name}参数",
                "example": self._generate_example_value(param),
                "type": param.type,
                "required": param.required,
                "width": self._calculate_column_width(param)
            }
            columns_info.append(column_info)
        
        # 添加结果列
        result_column = {
            "name": "执行结果",
            "description": "工作流执行结果（自动填充）",
            "example": "执行完成后自动填充",
            "type": WorkflowParameterType.TEXT,
            "required": False,
            "width": 30
        }
        columns_info.append(result_column)
        
        return columns_info
    
    def _calculate_column_width(self, param: WorkflowParameter) -> int:
        """计算列宽"""
        # 基于参数名长度和类型计算合适的列宽
        name_length = len(param.name)
        desc_length = len(param.description or "")
        
        if param.type == WorkflowParameterType.JSON:
            return min(self.max_column_width, max(30, name_length * 2))
        elif param.type == WorkflowParameterType.TEXT:
            return min(self.max_column_width, max(25, name_length * 1.5, desc_length * 0.8))
        else:
            return min(self.max_column_width, max(self.min_column_width, name_length * 1.2))
    
    def _write_header_row(self, ws: Worksheet, columns_info: List[Dict], row: int):
        """写入表头行"""
        for col, column_info in enumerate(columns_info, 1):
            cell = ws.cell(row=row, column=col, value=column_info["name"])
            
            # 应用样式
            cell.font = self.header_style["font"]
            cell.fill = self.header_style["fill"]
            cell.alignment = self.header_style["alignment"]
            cell.border = self.header_style["border"]
            
            # 必填字段标记
            if column_info.get("required"):
                cell.value = f"{column_info['name']} *"
    
    def _write_description_row(self, ws: Worksheet, columns_info: List[Dict], row: int):
        """写入描述行"""
        for col, column_info in enumerate(columns_info, 1):
            cell = ws.cell(row=row, column=col, value=column_info["description"])
            
            # 应用样式
            cell.font = self.description_style["font"]
            cell.alignment = self.description_style["alignment"]
            cell.border = self.description_style["border"]
    
    def _write_example_row(self, ws: Worksheet, columns_info: List[Dict], row: int):
        """写入示例行"""
        for col, column_info in enumerate(columns_info, 1):
            cell = ws.cell(row=row, column=col, value=column_info["example"])
            
            # 应用样式
            cell.font = self.example_style["font"]
            cell.alignment = self.example_style["alignment"]
            cell.border = self.example_style["border"]
    
    def _adjust_column_widths(self, ws: Worksheet, columns_info: List[Dict]):
        """调整列宽"""
        for col, column_info in enumerate(columns_info, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = column_info["width"]
    
    def _generate_example_value(self, param: WorkflowParameter) -> str:
        """生成参数示例值"""
        if param.default_value:
            return str(param.default_value)
        
        examples = {
            WorkflowParameterType.TEXT: "示例文本内容",
            WorkflowParameterType.NUMBER: "123",
            WorkflowParameterType.BOOLEAN: "true",
            WorkflowParameterType.JSON: '{"key": "value"}',
            WorkflowParameterType.FILE: "文件路径或URL"
        }
        
        if param.type == WorkflowParameterType.SELECT and param.options:
            return param.options[0] if param.options else "选项1"
        
        return examples.get(param.type, "示例值")
    
    def _create_instructions_sheet(self, wb: Workbook, parameters: WorkflowParameters):
        """创建使用说明工作表"""
        ws = wb.create_sheet("使用说明")
        
        current_row = 1
        
        # 标题
        cell = ws.cell(row=current_row, column=1, value="Dify工作流批量执行模板使用说明")
        cell.font = Font(bold=True, size=16, color="2F4F4F")
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{current_row}:D{current_row}")
        current_row += 2
        
        # 工作流信息
        current_row = self._add_section_title(ws, current_row, "📋 工作流信息")
        
        info_data = [
            ["工作流名称:", parameters.workflow_name],
            ["工作流ID:", parameters.workflow_id],
            ["参数数量:", str(len(parameters.parameters))],
            ["生成时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for label, value in info_data:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        current_row += 1
        
        # 使用步骤
        current_row = self._add_section_title(ws, current_row, "📝 使用步骤")
        
        steps = [
            "1. 切换到'批量数据'工作表",
            "2. 第1行为参数名称（请勿修改）",
            "3. 第2行为参数描述说明",
            "4. 第3行为示例数据格式",
            "5. 从第4行开始填写实际数据",
            "6. 必填参数用 * 标记，请确保填写",
            "7. 保存文件并上传到系统",
            "8. 系统将自动执行工作流",
            "9. 执行结果将填充到'执行结果'列"
        ]
        
        for step in steps:
            ws.cell(row=current_row, column=1, value=step)
            current_row += 1
        
        current_row += 1
        
        # 注意事项
        current_row = self._add_section_title(ws, current_row, "⚠️ 注意事项")
        
        notes = [
            "• 请确保数据格式正确，避免执行失败",
            "• 大量数据处理可能需要较长时间",
            "• 建议先用少量数据测试",
            "• JSON格式参数请使用标准JSON语法",
            "• 布尔值请使用 true/false",
            "• 数字类型请使用纯数字格式"
        ]
        
        for note in notes:
            ws.cell(row=current_row, column=1, value=note)
            current_row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_parameters_sheet(self, wb: Workbook, parameters: WorkflowParameters):
        """创建参数详情工作表"""
        ws = wb.create_sheet("参数详情")
        
        current_row = 1
        
        # 标题
        cell = ws.cell(row=current_row, column=1, value="工作流参数详细说明")
        cell.font = Font(bold=True, size=16, color="2F4F4F")
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 2
        
        # 表头
        headers = ["参数名称", "类型", "必填", "描述", "默认值", "可选项"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        current_row += 1
        
        # 参数详情
        for param in parameters.parameters:
            row_data = [
                param.name,
                param.type.value,
                "是" if param.required else "否",
                param.description or "无描述",
                str(param.default_value) if param.default_value else "无",
                ", ".join(param.options) if param.options else "无"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                
                # 必填参数高亮
                if col == 3 and value == "是":
                    cell.font = Font(bold=True, color="FF0000")
            
            current_row += 1
        
        # 调整列宽
        column_widths = [20, 15, 10, 40, 20, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _add_section_title(self, ws: Worksheet, row: int, title: str) -> int:
        """添加章节标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.subtitle_style["font"]
        cell.alignment = self.subtitle_style["alignment"]
        return row + 1
    
    def generate_result_template(self, columns: List[str], sample_data: List[Dict] = None) -> io.BytesIO:
        """
        生成结果模板
        
        Args:
            columns: 列名列表
            sample_data: 示例数据（可选）
            
        Returns:
            Excel文件的字节流
        """
        try:
            logger.info("开始生成结果模板")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "执行结果"
            
            # 写入表头
            for col, column_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=column_name)
                cell.font = self.header_style["font"]
                cell.fill = self.header_style["fill"]
                cell.alignment = self.header_style["alignment"]
                cell.border = self.header_style["border"]
            
            # 写入示例数据
            if sample_data:
                for row_idx, row_data in enumerate(sample_data, 2):
                    for col, column_name in enumerate(columns, 1):
                        value = row_data.get(column_name, "")
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )
            
            # 调整列宽
            for col in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col)].width = self.default_column_width
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("结果模板生成完成")
            return output
            
        except Exception as e:
            logger.error(f"生成结果模板失败: {str(e)}")
            raise FileProcessingException(f"生成结果模板失败: {str(e)}") 