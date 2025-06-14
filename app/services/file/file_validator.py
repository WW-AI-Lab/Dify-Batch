"""
文件验证器
"""
import os
import mimetypes
from typing import List, Dict, Any, Optional
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.exceptions import FileProcessingException
from app.core.logging import get_logger

logger = get_logger(__name__)


class FileValidator:
    """文件验证器"""
    
    def __init__(self):
        """初始化文件验证器"""
        self.max_file_size = 200 * 1024 * 1024  # 200MB（支持大文件）
        self.supported_extensions = ['.xlsx', '.xls']
        self.supported_mime_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
        ]
        self.max_rows = 100000  # 最大行数限制（支持10万行数据）
        self.max_columns = 100  # 最大列数限制
    
    def validate_upload_file(self, file_path: str, filename: str = None) -> Dict[str, Any]:
        """
        验证上传的文件
        
        Args:
            file_path: 文件路径
            filename: 原始文件名（可选）
            
        Returns:
            验证结果字典
        """
        result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "file_info": {}
        }
        
        try:
            logger.info(f"开始验证文件: {file_path}")
            
            # 基础文件验证
            basic_validation = self._validate_basic_file(file_path, filename)
            result["errors"].extend(basic_validation["errors"])
            result["warnings"].extend(basic_validation["warnings"])
            result["file_info"].update(basic_validation["file_info"])
            
            if basic_validation["errors"]:
                result["valid"] = False
                return result
            
            # Excel文件结构验证
            excel_validation = self._validate_excel_structure(file_path)
            result["errors"].extend(excel_validation["errors"])
            result["warnings"].extend(excel_validation["warnings"])
            result["file_info"].update(excel_validation["file_info"])
            
            if excel_validation["errors"]:
                result["valid"] = False
                return result
            
            # 内容验证
            content_validation = self._validate_excel_content(file_path)
            result["errors"].extend(content_validation["errors"])
            result["warnings"].extend(content_validation["warnings"])
            result["file_info"].update(content_validation["file_info"])
            
            if content_validation["errors"]:
                result["valid"] = False
            
            logger.info(f"文件验证完成: {file_path}, 有效: {result['valid']}")
            return result
            
        except Exception as e:
            logger.error(f"文件验证失败: {str(e)}")
            result["valid"] = False
            result["errors"].append(f"文件验证失败: {str(e)}")
            return result
    
    def _validate_basic_file(self, file_path: str, filename: str = None) -> Dict[str, Any]:
        """基础文件验证"""
        result = {
            "errors": [],
            "warnings": [],
            "file_info": {}
        }
        
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                result["errors"].append("文件不存在")
                return result
            
            # 获取文件信息
            file_stat = os.stat(file_path)
            file_size = file_stat.st_size
            file_ext = Path(file_path).suffix.lower()
            
            result["file_info"].update({
                "size": file_size,
                "size_mb": round(file_size / 1024 / 1024, 2),
                "extension": file_ext,
                "filename": filename or os.path.basename(file_path)
            })
            
            # 检查文件大小
            if file_size == 0:
                result["errors"].append("文件为空")
                return result
            
            if file_size > self.max_file_size:
                result["errors"].append(
                    f"文件大小({result['file_info']['size_mb']}MB)超过限制"
                    f"({self.max_file_size / 1024 / 1024:.1f}MB)"
                )
                return result
            
            # 检查文件扩展名
            if file_ext not in self.supported_extensions:
                result["errors"].append(f"不支持的文件格式: {file_ext}")
                return result
            
            # 检查MIME类型
            mime_type, _ = mimetypes.guess_type(file_path)
            result["file_info"]["mime_type"] = mime_type
            
            if mime_type and mime_type not in self.supported_mime_types:
                result["warnings"].append(f"文件MIME类型可能不正确: {mime_type}")
            
            return result
            
        except Exception as e:
            result["errors"].append(f"基础文件验证失败: {str(e)}")
            return result
    
    def _validate_excel_structure(self, file_path: str) -> Dict[str, Any]:
        """Excel文件结构验证"""
        result = {
            "errors": [],
            "warnings": [],
            "file_info": {}
        }
        
        try:
            # 尝试加载Excel文件
            try:
                workbook = load_workbook(file_path, read_only=True)
            except Exception as e:
                result["errors"].append(f"无法打开Excel文件: {str(e)}")
                return result
            
            # 获取工作表信息
            sheet_names = workbook.sheetnames
            result["file_info"]["sheet_names"] = sheet_names
            result["file_info"]["sheet_count"] = len(sheet_names)
            
            # 检查是否有"批量数据"工作表
            if "批量数据" not in sheet_names:
                result["errors"].append("缺少'批量数据'工作表")
                return result
            
            # 检查工作表内容
            worksheet = workbook["批量数据"]
            
            # 获取数据范围
            if worksheet.max_row is None or worksheet.max_column is None:
                result["errors"].append("工作表为空")
                return result
            
            # 安全地获取数据范围
            max_row = int(worksheet.max_row) if worksheet.max_row else 0
            max_column = int(worksheet.max_column) if worksheet.max_column else 0
            
            # 构建数据范围字符串，避免EmptyCell错误
            if max_row > 0 and max_column > 0:
                from openpyxl.utils import get_column_letter
                end_cell = f"{get_column_letter(max_column)}{max_row}"
                data_range = f"A1:{end_cell}"
            else:
                data_range = "A1:A1"
            
            result["file_info"].update({
                "max_row": max_row,
                "max_column": max_column,
                "data_range": data_range
            })
            
            # 检查行数和列数限制
            if worksheet.max_row > self.max_rows:
                result["errors"].append(f"数据行数({worksheet.max_row})超过限制({self.max_rows})")
            
            if worksheet.max_column > self.max_columns:
                result["errors"].append(f"数据列数({worksheet.max_column})超过限制({self.max_columns})")
            
            # 检查是否有足够的数据行（至少4行：表头、描述、示例、数据）
            if worksheet.max_row < 4:
                result["warnings"].append("数据行数较少，请确保至少有一行实际数据")
            
            workbook.close()
            return result
            
        except Exception as e:
            result["errors"].append(f"Excel结构验证失败: {str(e)}")
            return result
    
    def _validate_excel_content(self, file_path: str) -> Dict[str, Any]:
        """Excel内容验证"""
        result = {
            "errors": [],
            "warnings": [],
            "file_info": {}
        }
        
        try:
            # 使用pandas读取Excel文件（优化大文件处理）
            try:
                logger.info(f"正在读取Excel内容，可能需要一些时间...")
                df = pd.read_excel(file_path, sheet_name="批量数据", header=0, engine='openpyxl')
                logger.info(f"Excel内容读取完成，共{len(df)}行数据")
            except Exception as e:
                result["errors"].append(f"无法读取Excel内容: {str(e)}")
                return result
            
            # 获取列信息
            columns = df.columns.tolist()
            result["file_info"]["columns"] = columns
            result["file_info"]["column_count"] = len(columns)
            
            # 检查是否有列
            if not columns:
                result["errors"].append("Excel文件没有列标题")
                return result
            
            # 检查空列名
            empty_columns = [i for i, col in enumerate(columns) if pd.isna(col) or str(col).strip() == ""]
            if empty_columns:
                result["warnings"].append(f"发现空列名，位置: {empty_columns}")
            
            # 检查重复列名
            duplicate_columns = []
            seen_columns = set()
            for col in columns:
                if col in seen_columns:
                    duplicate_columns.append(col)
                seen_columns.add(col)
            
            if duplicate_columns:
                result["errors"].append(f"发现重复列名: {duplicate_columns}")
            
            # 过滤空行
            df_filtered = df.dropna(how='all')
            result["file_info"]["total_rows"] = int(len(df))
            result["file_info"]["non_empty_rows"] = int(len(df_filtered))
            
            # 跳过描述行和示例行，计算实际数据行
            if len(df_filtered) > 2:
                data_df = df_filtered.iloc[2:]  # 从第3行开始
                actual_data_rows = len(data_df.dropna(how='all'))
                result["file_info"]["data_rows"] = int(actual_data_rows)
                
                if actual_data_rows == 0:
                    result["warnings"].append("没有找到实际数据行")
                elif actual_data_rows > 50000:
                    result["warnings"].append(f"数据行数很多({actual_data_rows})，建议分批处理或降低并发数")
                elif actual_data_rows > 10000:
                    result["warnings"].append(f"数据行数较多({actual_data_rows})，处理可能需要较长时间")
            else:
                result["file_info"]["data_rows"] = 0
                result["warnings"].append("没有找到实际数据行")
            
            # 检查数据完整性
            self._check_data_completeness(df, result)
            
            return result
            
        except Exception as e:
            result["errors"].append(f"Excel内容验证失败: {str(e)}")
            return result
    
    def _check_data_completeness(self, df: pd.DataFrame, result: Dict[str, Any]):
        """检查数据完整性"""
        try:
            # 跳过描述行和示例行
            if len(df) > 2:
                data_df = df.iloc[2:]
                
                # 统计每列的空值情况
                null_stats = {}
                for col in df.columns:
                    if col != "执行结果":  # 排除结果列
                        null_count = data_df[col].isna().sum()
                        total_count = len(data_df)
                        null_percentage = (null_count / total_count * 100) if total_count > 0 else 0
                        
                        null_stats[col] = {
                            "null_count": int(null_count),  # 转换为Python int
                            "total_count": int(total_count),  # 转换为Python int
                            "null_percentage": round(float(null_percentage), 2)  # 转换为Python float
                        }
                        
                        # 如果某列空值过多，给出警告
                        if null_percentage > 50:
                            result["warnings"].append(
                                f"列'{col}'空值较多({null_percentage:.1f}%)，请检查数据完整性"
                            )
                
                result["file_info"]["null_statistics"] = null_stats
                
        except Exception as e:
            result["warnings"].append(f"数据完整性检查失败: {str(e)}")
    
    def validate_template_format(self, file_path: str, expected_columns: List[str]) -> Dict[str, Any]:
        """
        验证模板格式是否符合预期
        
        Args:
            file_path: 文件路径
            expected_columns: 期望的列名列表
            
        Returns:
            验证结果
        """
        result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "column_mapping": {}
        }
        
        try:
            logger.info(f"验证模板格式: {file_path}")
            
            # 读取Excel文件
            df = pd.read_excel(file_path, sheet_name="批量数据", header=0)
            actual_columns = [col for col in df.columns.tolist() if col != "执行结果"]
            
            # 检查列名匹配
            missing_columns = []
            extra_columns = []
            
            for expected_col in expected_columns:
                if expected_col not in actual_columns:
                    missing_columns.append(expected_col)
            
            for actual_col in actual_columns:
                if actual_col not in expected_columns:
                    extra_columns.append(actual_col)
            
            # 生成列映射
            for col in expected_columns:
                if col in actual_columns:
                    result["column_mapping"][col] = col
            
            # 处理验证结果
            if missing_columns:
                result["valid"] = False
                result["errors"].append(f"缺少必需的列: {missing_columns}")
            
            if extra_columns:
                result["warnings"].append(f"发现额外的列: {extra_columns}")
            
            result["file_info"] = {
                "expected_columns": expected_columns,
                "actual_columns": actual_columns,
                "missing_columns": missing_columns,
                "extra_columns": extra_columns
            }
            
            logger.info(f"模板格式验证完成: {result['valid']}")
            return result
            
        except Exception as e:
            logger.error(f"模板格式验证失败: {str(e)}")
            result["valid"] = False
            result["errors"].append(f"模板格式验证失败: {str(e)}")
            return result 