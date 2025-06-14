#!/usr/bin/env python3
"""
创建测试Excel文件
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建测试Excel文件
wb = Workbook()
ws = wb.active
ws.title = '批量数据'

# 设置样式
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 写入表头
headers = ['query *', '执行结果']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# 写入描述行
descriptions = ['搜索词', '工作流执行结果（自动填充）']
for col, desc in enumerate(descriptions, 1):
    cell = ws.cell(row=2, column=col, value=desc)
    cell.border = border

# 写入示例行
examples = ['示例文本内容', '执行完成后自动填充']
for col, example in enumerate(examples, 1):
    cell = ws.cell(row=3, column=col, value=example)
    cell.border = border

# 写入测试数据
test_data = [
    ['Python编程教程', ''],
    ['机器学习入门', ''],
    ['Web开发指南', '']
]

for row_idx, row_data in enumerate(test_data, 4):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = border

# 调整列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30

# 保存文件
wb.save('test_upload_file.xlsx')
print('测试文件创建成功: test_upload_file.xlsx') 